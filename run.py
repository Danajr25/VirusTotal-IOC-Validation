import requests
import pandas as pd
from openpyxl import load_workbook
import re
import ipaddress

# === CONFIGURATION ===
VT_API_KEY = 'API_KEY'  # <-- Replace this with your actual API key
FILE1_PATH = 'File1.xlsx'
FILE2_PATH = 'File2.xlsx'

def vt_lookup(hash_value):
    url = f"https://www.virustotal.com/api/v3/files/{hash_value}"
    headers = {"x-apikey": VT_API_KEY}
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        data = r.json()
        attributes = data.get("data", {}).get("attributes", {})
        analysis = attributes.get("last_analysis_results", {})
        
        # Use the correct engine name: "Paloalto" (not "Palo Alto Networks")
        palo_alto_result = analysis.get("Paloalto", {})
        
        # Try both 'category' and 'result' fields
        verdict = palo_alto_result.get("category", palo_alto_result.get("result", "Not Found"))
        detected = attributes.get("last_analysis_stats", {}).get("malicious", 0) > 0

        return {
            "sha256": attributes.get("sha256", "N/A"),
            "sha1": attributes.get("sha1", "N/A"),
            "md5": attributes.get("md5", "N/A"),
            "detected": detected,
            "palo_verdict": verdict
        }
    elif r.status_code == 404:
        return {"not_found": True}
    else:
        return {"error": True, "code": r.status_code}

def vt_ip_lookup(ip_address):
    url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip_address}"
    headers = {"x-apikey": VT_API_KEY}
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        data = r.json()
        attributes = data.get("data", {}).get("attributes", {})
        analysis = attributes.get("last_analysis_results", {})
        
        # Use the correct engine name: "Paloalto" (not "Palo Alto Networks")
        palo_alto_result = analysis.get("Paloalto", {})
        
        # Try both 'category' and 'result' fields
        verdict = palo_alto_result.get("category", palo_alto_result.get("result", "Not Found"))
        detected = attributes.get("last_analysis_stats", {}).get("malicious", 0) > 0

        return {
            "ip": ip_address,
            "detected": detected,
            "palo_verdict": verdict
        }
    elif r.status_code == 404:
        return {"not_found": True}
    else:
        return {"error": True, "code": r.status_code}

def is_cve(value):
    # Check if it's a valid CVE format (CVE-YYYY-NNNN)
    return re.match(r'^CVE-\d{4}-\d{4,}$', value, re.IGNORECASE) is not None

def is_ip_address(value):
    try:
        ipaddress.ip_address(value)
        return True
    except ValueError:
        return False

def is_hash(value):
    # Check if it's a valid hash (MD5: 32 chars, SHA1: 40 chars, SHA256: 64 chars)
    if re.match(r'^[a-fA-F0-9]{32}$', value):  # MD5
        return True
    elif re.match(r'^[a-fA-F0-9]{40}$', value):  # SHA1
        return True
    elif re.match(r'^[a-fA-F0-9]{64}$', value):  # SHA256
        return True
    return False

def insert_into_table(ws, row_data):
    # Find the table headers (MD5, SHA1, SHA256) in the sheet
    header_row = None
    for row in range(1, ws.max_row + 1):
        if (ws[f"B{row}"].value == "MD5" and 
            ws[f"C{row}"].value == "SHA1" and 
            ws[f"D{row}"].value == "SHA256"):
            header_row = row
            break
    
    if header_row is None:
        print("Warning: Could not find table headers (MD5, SHA1, SHA256) in the sheet")
        return
    
    # Find the first truly empty row after the headers
    data_start_row = header_row + 1
    row = data_start_row
    while row <= ws.max_row:
        # Check if the SHA256 column (D) has any content (including N/A)
        sha256_cell = ws[f"D{row}"].value
        if sha256_cell is None or sha256_cell == "":
            break
        row += 1
    
    # Insert data into the table
    ws[f"B{row}"] = row_data["md5"]
    ws[f"C{row}"] = row_data["sha1"] 
    ws[f"D{row}"] = row_data["sha256"]
    print(f"    Inserted hash data into row {row}")

def insert_into_ip_table(ws, ip_data):
    # Find the next empty row in the DomainIP column (B) starting from row 2
    next_row = 2
    
    while next_row <= ws.max_row:
        domainip_cell = ws[f"B{next_row}"].value
        if domainip_cell is None or domainip_cell == "" or domainip_cell == "N/A":
            break
        next_row += 1
    
    # Fill the IP and verdict
    ws[f"B{next_row}"] = ip_data["ip"]
    
    # Determine verdict text
    if "not_found" in ip_data:
        verdict_text = "Not Found"
    elif ip_data["detected"]:
        if ip_data["palo_verdict"] in ["malicious"]:
            verdict_text = "Malware Covered by XDR"
        else:
            verdict_text = "Malware Not Covered by XDR"
    else:
        verdict_text = "Clean"
    
    ws[f"C{next_row}"] = verdict_text
    print(f"    Inserted IP data into row {next_row}")

def insert_into_cve_table(ws, cve_value):
    # Find the next empty row in the CVE column (B) starting from row 2
    next_row = 2
    
    while next_row <= ws.max_row:
        cve_cell = ws[f"B{next_row}"].value
        if cve_cell is None or cve_cell == "" or cve_cell == "N/A":
            break
        next_row += 1
    
    # Fill the CVE value
    ws[f"B{next_row}"] = cve_value
    print(f"    Inserted CVE into row {next_row}")

def main():
    df = pd.read_excel(FILE2_PATH, header=None)
    entries = df[0].dropna().tolist()

    wb = load_workbook(FILE1_PATH)
    
    # Access the sheets
    sheet1 = wb['Sheet1']  # Malware covered by XDR (Palo Alto detects as malicious)
    sheet2 = wb['Sheet2']  # Not found or clean
    sheet3 = wb['Sheet3']  # Malware not covered by XDR (Palo Alto doesn't detect as malicious)
    domainip_sheet = wb['DomainIP']  # IP addresses go here
    cve_sheet = wb['CVE']  # CVEs go here

    for entry in entries:
        entry = str(entry).strip()
        print(f"Checking: {entry}")
        
        if is_ip_address(entry):
            # Handle IP address - all IPs go to DomainIP sheet
            result = vt_ip_lookup(entry)
            
            if "error" in result:
                print(f"Error: {result['code']} for IP {entry}")
                continue
            if "not_found" in result:
                # IP not found on VirusTotal
                insert_into_ip_table(domainip_sheet, {"ip": entry, "not_found": True})
                print(f"  → DomainIP: IP not found on VirusTotal")
                continue

            if result["detected"]:
                verdict = result["palo_verdict"]
                if verdict in ["malicious"]:
                    # Malicious IP AND Palo Alto identifies it as malicious
                    insert_into_ip_table(domainip_sheet, result)
                    print(f"  → DomainIP: Malicious IP covered by XDR (Palo Alto verdict: {verdict})")
                else:
                    # Malicious IP BUT Palo Alto doesn't identify it as malicious
                    insert_into_ip_table(domainip_sheet, result)
                    print(f"  → DomainIP: Malicious IP not covered by XDR (Palo Alto verdict: {verdict})")
            else:
                # Clean IP
                insert_into_ip_table(domainip_sheet, result)
                print(f"  → DomainIP: Clean IP")
                
        elif is_cve(entry):
            # Handle CVE - just paste it to CVE sheet, no VirusTotal check needed
            insert_into_cve_table(cve_sheet, entry)
            print(f"  → CVE: Added to CVE sheet")
                
        elif is_hash(entry):
            # Handle valid hash
            result = vt_lookup(entry)

            if "error" in result:
                print(f"Error: {result['code']} for hash {entry}")
                continue
            if "not_found" in result:
                # Hash not found on VirusTotal - goes to Sheet2
                insert_into_table(sheet2, {"md5": "N/A", "sha1": "N/A", "sha256": entry})
                print(f"  → Sheet2: Hash not found on VirusTotal")
                continue

            if result["detected"]:
                verdict = result["palo_verdict"]
                if verdict in ["malicious"]:
                    # Malware detected AND Palo Alto identifies it as malicious - goes to Sheet1 (covered by XDR)
                    insert_into_table(sheet1, result)
                    print(f"  → Sheet1: Malware covered by XDR (Palo Alto verdict: {verdict})")
                else:
                    # Malware detected BUT Palo Alto doesn't identify it as malicious - goes to Sheet3 (not covered by XDR)
                    insert_into_table(sheet3, result)
                    print(f"  → Sheet3: Malware not covered by XDR (Palo Alto verdict: {verdict})")
            else:
                # Clean hash (not detected as malware) - goes to Sheet2
                insert_into_table(sheet2, result)
                print(f"  → Sheet2: Clean hash")
        else:
            # Invalid entry - treat as "not found" hash and put in Sheet2
            insert_into_table(sheet2, {"md5": "N/A", "sha1": "N/A", "sha256": entry})
            print(f"  → Sheet2: Invalid entry treated as not found hash")

    wb.save("File1_filled.xlsx")
    print("✅ Done. Output saved to File1_filled.xlsx")

if __name__ == "__main__":
    main()
